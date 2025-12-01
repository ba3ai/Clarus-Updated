from __future__ import annotations

from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
import os, math, traceback
from datetime import datetime, timedelta, date
from calendar import monthrange
from typing import Dict, Any, List, Tuple, Optional

from openpyxl import load_workbook
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
import re  # vanilla re (used for filename year detection etc.)

from backend.models import (
    db,
    ExcelUploadHistory,
    PortfolioPeriodMetric,   # still imported for compatibility
    Investment,
    PortfolioInvestmentValue,
    AdminPeriodBalance,      # NEW admin totals table
)
from backend.models_snapshot import InvestorPeriodBalance  # NEW investor-period table

# Optional lineage record
try:
    from backend.models import DataSource
except Exception:
    DataSource = None  # pragma: no cover

# Optional current_user (if your app uses flask-login)
try:
    from flask_login import current_user  # type: ignore
except Exception:
    current_user = None  # pragma: no cover

excel_bp = Blueprint("excel_bp", __name__, url_prefix="/excel")

UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

DATE_COL = "as_of_date"


# ---------------------- generic helpers ----------------------
def _dialect_insert():
    name = (db.engine.name or "").lower()
    return pg_insert if "postgre" in name else sqlite_insert


def _to_float_cell(v) -> float:
    if v is None or v == "" or str(v).strip() in {"—", "-", "–"}:
        return math.nan
    s = str(v).strip().replace(",", "").replace("$", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except Exception:
        return math.nan


def _candidate_date_formats():
    return (
        "%m/%d/%Y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%y",
        "%d-%b-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%d-%b-%y",
        "%b-%y",
        "%b %y",
        "%b-%Y",
        "%b %Y",
    )


def _maybe_excel_serial(v) -> Optional[date]:
    try:
        fv = float(v)
        if 20000 < fv < 90000:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(fv))).date()
    except Exception:
        pass
    return None


def _looks_like_date(v) -> bool:
    if isinstance(v, (datetime, date)):
        return True
    if _maybe_excel_serial(v):
        return True
    s = str(v).strip().rstrip("Z")
    for fmt in _candidate_date_formats():
        try:
            datetime.strptime(s, fmt)
            return True
        except Exception:
            continue
    return False


def _parse_date_any(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    ser = _maybe_excel_serial(v)
    if ser:
        return ser
    s = str(v).strip().rstrip("Z")
    for fmt in _candidate_date_formats():
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


import regex as _re  # more powerful regex


def _clean_txt(x: str) -> str:
    return _re.sub(r"\s+", " ", (x or "")).strip().lower()


def _values_from_openpyxl(ws) -> list[list]:
    vals = []
    for r in ws.iter_rows(values_only=True):
        vals.append(list(r))
    return vals


# ---------------------- NEW BALANCE-SHEET INGEST (tidy format) ----------------------
def _normalize_header_key(raw: str) -> str:
    """
    Normalize header text to a simple key:
    - lowercased
    - spaces, dashes, punctuation -> underscores
    - strip leading/trailing underscores
    """
    txt = _clean_txt(raw)
    txt = _re.sub(r"[^a-z0-9]+", "_", txt)
    return txt.strip("_")


# Mapping normalized header -> DB field on InvestorPeriodBalance
_BALANCE_HEADER_MAP = {
    # investor name
    "name": "name",
    "investor": "name",
    "investor_name": "name",
    "partner_name": "name",

    # as-of / period (optional; we will derive from sheet if missing)
    "as_of_date": "as_of_date",
    "as_of": "as_of_date",
    "period": "as_of_date",
    "period_end": "as_of_date",
    "period_end_date": "as_of_date",

    # metrics
    "beginning_ownership": "beginning_ownership",
    "beginning_balance": "beginning_balance",
    "beginning_nav": "beginning_balance",
    "opening_balance": "beginning_balance",
    "gross_profit": "gross_profit",
    "gross_p_l": "gross_profit",
    "management_fees": "management_fees",
    "mgmt_fees": "management_fees",
    "operating_expenses": "operating_expenses",
    "operating_expense": "operating_expenses",
    "allocated_fees": "allocated_fees",
    "additions": "additions",
    "withdrawals": "withdrawals",
    "withdrawls": "withdrawals",  # allow common typo
    "ending_balance": "ending_balance",
    "closing_balance": "ending_balance",
    "ending_nav": "ending_balance",
}


def _find_balance_header_row(values: List[List]) -> Tuple[Optional[int], Dict[str, int]]:
    """
    Find the header row for the NEW balance-sheet format and return:
      (row_index_0_based, {db_field_name -> col_index_0_based})

    For Clarus sheets, we ONLY require a 'name' column.
    If an 'as_of_date' column is not present, we will derive the date from the sheet name.
    """
    max_scan = min(50, len(values))
    for i in range(max_scan):
        row = values[i] or []
        mapping: Dict[str, int] = {}
        for j, cell in enumerate(row):
            if cell is None:
                continue
            norm = _normalize_header_key(str(cell))
            field = _BALANCE_HEADER_MAP.get(norm)
            if field:
                mapping.setdefault(field, j)

        # we require at least a name column to treat this as the header row
        if "name" in mapping:
            return i, mapping

    return None, {}


_MONTH_WORDS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def _derive_as_of_date_from_sheet(sheet_name: str) -> Optional[date]:
    """
    Try to infer a month-end date from the sheet name, e.g.:
      'November-2025' -> 2025-11-30
      'Nov 2025'      -> 2025-11-30
      '2025-11'       -> 2025-11-30

    Returns None if it can't parse.
    """
    if not sheet_name:
        return None

    s = sheet_name.strip().lower()
    s = _re.sub(r"[_\.\(\)]", " ", s)

    # 1) Look for month word + 4-digit year
    tokens = [t for t in _re.split(r"[^a-z0-9]+", s) if t]
    month = None
    year = None
    for t in tokens:
        if t in _MONTH_WORDS and month is None:
            month = _MONTH_WORDS[t]
        elif _re.fullmatch(r"20\d{2}", t) and year is None:
            year = int(t)

    if month and year:
        return date(year, month, monthrange(year, month)[1])

    # 2) Fallback: pure numeric year-month like 2025-11, 2025_11, etc.
    m = _re.search(r"(20\d{2})\D+(\d{1,2})", s)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        if 1 <= month <= 12:
            return date(year, month, monthrange(year, month)[1])

    return None


def _ingest_new_balance_sheet(values: List[List], resolved_sheet: str) -> Dict[str, Any]:
    """
    Ingest the NEW tidy balance sheet format into:
      - InvestorPeriodBalance (per investor, per as_of_date)
      - AdminPeriodBalance (per as_of_date)

    Admin totals:
      - If a row named 'Total' / 'Grand Total' exists for a period, we use THAT row
        for AdminPeriodBalance (source='total_row').
      - Otherwise, we fall back to summing all investor rows
        (source='aggregated').

    For Clarus:
      - Header row has 'Partner name', 'Beginning Balance', 'Ending Balance', etc.
      - Sheet name is like 'November-2025' and we derive as_of_date from that
        when there is no explicit As Of Date column.
    """
    header_idx, col_map = _find_balance_header_row(values)
    if header_idx is None or not col_map:
        raise RuntimeError(
            "Could not find header row for new balance sheet format "
            "(expected at least a 'Name' / 'Partner name' column)."
        )

    body = values[header_idx + 1 :]

    insert_fn = _dialect_insert()
    investor_rows = 0

    # Determine whether we have an as_of_date column, or need to derive from sheet name
    has_asof_col = "as_of_date" in col_map
    sheet_as_of: Optional[date] = None
    if not has_asof_col:
        sheet_as_of = _derive_as_of_date_from_sheet(resolved_sheet)
        if not sheet_as_of:
            raise RuntimeError(
                "Could not determine period date: no 'As Of Date' column and "
                f"sheet name '{resolved_sheet}' does not contain a recognizable month/year "
                "(e.g. 'November-2025')."
            )

    metric_fields = [
        "beginning_ownership",
        "beginning_balance",
        "gross_profit",
        "management_fees",
        "operating_expenses",
        "allocated_fees",
        "additions",
        "withdrawals",
        "ending_balance",
    ]

    # For fallback aggregation (non-total investor rows)
    admin_acc: Dict[date, Dict[str, float]] = {}
    admin_seen: Dict[date, Dict[str, bool]] = {}

    # For explicit 'Total' rows (one per as_of_date)
    total_rows: Dict[date, Dict[str, Optional[float]]] = {}

    blank_streak = 0
    for r in body:
        # --- investor / total name ---
        name_idx = col_map["name"]
        raw_name = r[name_idx] if name_idx < len(r) else None
        name = str(raw_name or "").strip()
        if not name:
            blank_streak += 1
            if blank_streak > 50:
                break
            continue
        blank_streak = 0

        clean_name = _clean_txt(name)

        # --- as_of_date ---
        if has_asof_col:
            asof_idx = col_map["as_of_date"]
            raw_date = r[asof_idx] if asof_idx < len(r) else None
            as_of = _parse_date_any(raw_date)
            if not as_of:
                # Skip rows without a valid date if we *expect* a column
                continue
        else:
            # Clarus-style: same as_of_date for all rows in this sheet
            as_of = sheet_as_of

        row_data: Dict[str, Any] = {"name": name, "as_of_date": as_of}

        # metrics
        for field in metric_fields:
            if field not in col_map:
                row_data[field] = None
                continue
            idx = col_map[field]
            val = r[idx] if idx < len(r) else None
            f = _to_float_cell(val)
            row_data[field] = None if math.isnan(f) else float(f)

        # ---------- TOTAL ROW LOGIC (for admin dashboard) ----------
        if clean_name in {"total", "grand total"}:
            # Capture totals for this period. We don't insert an investor row
            # and we don't use this row for aggregation.
            totals = total_rows.setdefault(
                as_of, {m: None for m in metric_fields}
            )
            for m in metric_fields:
                v = row_data.get(m)
                if v is not None:
                    totals[m] = float(v)
            # Skip InvestorPeriodBalance upsert for the 'Total' row
            continue
        # -----------------------------------------------------------

        # --- upsert InvestorPeriodBalance for NORMAL investor rows ---
        stmt = insert_fn(InvestorPeriodBalance).values(
            name=row_data["name"],
            as_of_date=row_data["as_of_date"],
            beginning_ownership=row_data["beginning_ownership"],
            beginning_balance=row_data["beginning_balance"],
            gross_profit=row_data["gross_profit"],
            management_fees=row_data["management_fees"],
            operating_expenses=row_data["operating_expenses"],
            allocated_fees=row_data["allocated_fees"],
            additions=row_data["additions"],
            withdrawals=row_data["withdrawals"],
            ending_balance=row_data["ending_balance"],
            source="excel-upload",
        )

        stmt = stmt.on_conflict_do_update(
            index_elements=["name", "as_of_date"],
            set_={
                "beginning_ownership": stmt.excluded.beginning_ownership,
                "beginning_balance": stmt.excluded.beginning_balance,
                "gross_profit": stmt.excluded.gross_profit,
                "management_fees": stmt.excluded.management_fees,
                "operating_expenses": stmt.excluded.operating_expenses,
                "allocated_fees": stmt.excluded.allocated_fees,
                "additions": stmt.excluded.additions,
                "withdrawals": stmt.excluded.withdrawals,
                "ending_balance": stmt.excluded.ending_balance,
                "source": stmt.excluded.source,
                "updated_at": datetime.utcnow(),
            },
        )
        db.session.execute(stmt)
        investor_rows += 1

        # --- fallback aggregation for admin totals (non-total rows only) ---
        acc = admin_acc.setdefault(as_of, {m: 0.0 for m in metric_fields})
        seen = admin_seen.setdefault(as_of, {m: False for m in metric_fields})
        for m in metric_fields:
            v = row_data.get(m)
            if v is None:
                continue
            acc[m] += float(v)
            seen[m] = True

    # Commit investor rows first
    db.session.commit()

    # --- upsert AdminPeriodBalance per as_of_date ---
    admin_periods: List[str] = []

    # We want union of all periods that either have totals or aggregated values
    all_periods = sorted(set(admin_acc.keys()) | set(total_rows.keys()))

    for as_of in all_periods:
        payload: Dict[str, Any] = {"as_of_date": as_of}

        if as_of in total_rows:
            # Prefer explicit Total row for this period
            totals = total_rows[as_of]
            for m in metric_fields:
                v = totals.get(m)
                payload[m] = v if v is not None else None
            payload["source"] = "total_row"
        else:
            # Fallback: aggregated from investor rows
            sums = admin_acc.get(as_of, {m: 0.0 for m in metric_fields})
            seen = admin_seen.get(as_of, {m: False for m in metric_fields})
            payload["beginning_ownership"] = (
                sums["beginning_ownership"] if seen["beginning_ownership"] else None
            )
            payload["beginning_balance"] = (
                sums["beginning_balance"] if seen["beginning_balance"] else None
            )
            payload["gross_profit"] = sums["gross_profit"] if seen["gross_profit"] else None
            payload["management_fees"] = (
                sums["management_fees"] if seen["management_fees"] else None
            )
            payload["operating_expenses"] = (
                sums["operating_expenses"] if seen["operating_expenses"] else None
            )
            payload["allocated_fees"] = (
                sums["allocated_fees"] if seen["allocated_fees"] else None
            )
            payload["additions"] = sums["additions"] if seen["additions"] else None
            payload["withdrawals"] = sums["withdrawals"] if seen["withdrawals"] else None
            payload["ending_balance"] = (
                sums["ending_balance"] if seen["ending_balance"] else None
            )
            payload["source"] = "aggregated"

        stmt = insert_fn(AdminPeriodBalance).values(**payload)
        stmt = stmt.on_conflict_do_update(
            index_elements=["as_of_date"],
            set_={
                "beginning_ownership": stmt.excluded.beginning_ownership,
                "beginning_balance": stmt.excluded.beginning_balance,
                "gross_profit": stmt.excluded.gross_profit,
                "management_fees": stmt.excluded.management_fees,
                "operating_expenses": stmt.excluded.operating_expenses,
                "allocated_fees": stmt.excluded.allocated_fees,
                "additions": stmt.excluded.additions,
                "withdrawals": stmt.excluded.withdrawals,
                "ending_balance": stmt.excluded.ending_balance,
                "source": stmt.excluded.source,
                "updated_at": datetime.utcnow(),
            },
        )
        db.session.execute(stmt)
        admin_periods.append(as_of.isoformat())

    db.session.commit()

    return {
        "ok": True,
        "sheet": resolved_sheet,
        "investor_period_rows": investor_rows,
        "admin_periods_upserted": admin_periods,
    }


# ---------------------- SHEET RESOLUTION (mirror SP) ----------------------
import re as _re2


def _normalize_sheet_name(s: str) -> str:
    return _re2.sub(r"[\s\-\_\.\(\)\[\]\{\}\+]+", "", (s or "").strip()).lower()


def _sheet_candidates(name: str) -> list[str]:
    base = (name or "").strip()
    if not base:
        return []
    variants = {
        base,
        base.replace("(", " (").replace("  ", " ").strip(),
        _re2.sub(r"\s*\(", " (", base).strip(),
        _re2.sub(r"\s+", "", base),
        base.replace("+", " "),
    }
    return [v for v in variants if v]


# ---------------------- FILE CLASSIFIER ----------------------
def _has_investments_table(values: List[List]) -> bool:
    """
    Detect whether a sheet looks like the Investments grid.

    In your new file the header is 'Investment' (singular), but we also
    accept 'Investments' for backwards compatibility.
    """
    limit = min(200, len(values))
    for r in range(limit):
        row = values[r] or []
        for cell in row:
            txt = str(cell or "").strip().lower()
            if txt in {"investment", "investments"}:
                return True
    return False


def _has_balance_labels(values: List[List]) -> bool:
    # For new format we still expect at least "Beginning Balance" and maybe "Ending Balance" in header.
    labels = {
        "beginning balance",
        "ending balance",
        "gross profit",
        "management fees",
        "operating expenses",
        "allocated fees",
        "additions",
        "withdrawals",
    }
    limit = min(250, len(values))
    for r in range(limit):
        row = values[r] or []
        for cell in row:
            txt = str(cell or "").strip().lower()
            if txt in labels:
                return True
    return False


def _classify_workbook(values: List[List]) -> str:
    has_inv = _has_investments_table(values)
    has_bal = _has_balance_labels(values)
    if has_inv and has_bal:
        return "mixed"
    if has_inv:
        return "investment"
    if has_bal:
        return "balance"
    return "unknown"


# ---------------------- INVESTMENTS INGEST ----------------------
_MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def _find_header_row(values: List[List]) -> int:
    upto = min(80, len(values))
    for i in range(upto):
        row = [str(x or "").strip().lower() for x in values[i]]
        if any(_re.match(r"^invest(ment|ments)\b", c) for c in row):
            return i
    for i in range(upto):
        row = [str(x or "").strip().lower() for x in values[i]]
        if any(row):
            return i
    return 0


def _month_end(d: date) -> date:
    return date(d.year, d.month, monthrange(d.year, d.month)[1])


def _detect_year_banners(values: List[List], top_rows: int = 6) -> Dict[int, int]:
    year_by_col: Dict[int, int] = {}
    upto = min(top_rows, len(values))
    for r in range(upto):
        row = values[r] or []
        for j, cell in enumerate(row):
            if isinstance(cell, (int, float)):
                continue
            txt_raw = "" if cell is None else str(cell)
            txt = _clean_txt(txt_raw)
            if not txt:
                continue
            if _re.fullmatch(r"\s*(20\d{2})\s*", txt):
                year_by_col.setdefault(j, int(_re.fullmatch(r"\s*(20\d{2})\s*", txt).group(1)))
                continue
            if not _re.search(r"[a-z]", txt):
                continue
            m = _re.search(r"\b(20\d{2})\b", txt)
            if m:
                year_by_col.setdefault(j, int(m.group(1)))
    return year_by_col


def _detect_date_columns(
    values: List[List],
    header_row_idx: int,
    preferred_year: Optional[int] = None,
) -> Dict[int, date]:
    from calendar import monthrange as _mr

    rows = len(values)
    cols = max((len(r) for r in values), default=0)

    def _row_txts(ridx: int) -> List[str]:
        row = values[ridx] if 0 <= ridx < rows else []
        return [_clean_txt(str(row[j])) if j < len(row) else "" for j in range(cols)]

    MONTHS = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }

    def _month_map_for_row(ridx: int) -> Dict[int, date]:
        out: Dict[int, date] = {}
        txts = _row_txts(ridx)
        for j, t in enumerate(txts):
            if t in MONTHS:
                m = MONTHS[t]
                y = preferred_year if preferred_year is not None else None
                if y:
                    out[j] = date(y, m, _mr(y, m)[1])
        return out

    if preferred_year is not None:
        out = _month_map_for_row(header_row_idx)
        if out:
            return out
        for off in [1, -1, 2, -2, 3, -3]:
            out = _month_map_for_row(header_row_idx + off)
            if out:
                return out
        for ridx in range(min(rows, 8)):
            out = _month_map_for_row(ridx)
            if out:
                return out

    for r in range(0, min(rows, 120)):
        row = values[r] if r < rows else []
        if any(_clean_txt(str(c)) == "ending date" for c in row):
            out: Dict[int, date] = {}
            for j, c in enumerate(row):
                dt = _parse_date_any(c)
                if dt:
                    out[j] = date(dt.year, dt.month, _mr(dt.year, dt.month)[1])
            if out:
                return out

    best_r = None
    best_count = 0
    for r in range(0, min(rows, 120)):
        row = values[r] if r < rows else []
        cnt = sum(1 for c in row if _parse_date_any(c))
        if cnt > best_count:
            best_count, best_r = cnt, r
    if best_r is not None and best_count >= 4:
        row = values[best_r]
        out: Dict[int, date] = {}
        for j, c in enumerate(row):
            dt = _parse_date_any(c)
            if dt:
                out[j] = date(dt.year, dt.month, _mr(dt.year, dt.month)[1])
        if out:
            return out

    header = values[header_row_idx] if header_row_idx < rows else []
    out: Dict[int, date] = {}
    for j in range(cols):
        dt = _parse_date_any(header[j] if j < len(header) else None)
        if dt:
            out[j] = date(dt.year, dt.month, _mr(dt.year, dt.month)[1])
    if out:
        return out

    year_banners: Dict[int, int] = _detect_year_banners(values, top_rows=6)

    def _nearest_banner_year_to_right(col: int) -> Optional[int]:
        rights = [c for c in sorted(year_banners) if c > col]
        return year_banners[rights[0]] if rights else None

    header_txts = _row_txts(header_row_idx)
    current_year: Optional[int] = None
    out = {}
    for j in range(cols):
        if j in year_banners:
            current_year = year_banners[j]
        hdr = header_txts[j]
        if hdr in MONTHS:
            y = current_year if current_year is not None else _nearest_banner_year_to_right(j)
            if y:
                m = MONTHS[hdr]
                out[j] = date(y, m, _mr(y, m)[1])
    if out:
        return out

    out = {}
    for j in range(cols):
        r = header_row_idx - 1
        while r >= 0:
            dt = _parse_date_any(values[r][j] if j < len(values[r]) else None)
            if dt:
                out[j] = date(dt.year, dt.month, _mr(dt.year, dt.month)[1])
                break
            r -= 1
    return out


def _ensure_color(i: int) -> str:
    palette = [
        "#6366F1",
        "#10B981",
        "#F59E0B",
        "#EC4899",
        "#14B8A6",
        "#F43F5E",
        "#22C55E",
        "#8B5CF6",
        "#06B6D4",
        "#F97316",
        "#84CC16",
        "#3B82F6",
        "#A855F7",
        "#0EA5E9",
        "#D946EF",
    ]
    return palette[i % len(palette)]


def _ingest_investments_table(
    values: List[List],
    resolved_sheet: str,
    source_id: Optional[int],
    preferred_year: Optional[int] = None,
) -> Dict[str, Any]:
    if not values or len(values) < 2:
        return {"ok": False, "error": "No cells to parse", "sheet": resolved_sheet}

    header_row_idx = _find_header_row(values)
    header = [str(c or "").strip() for c in values[header_row_idx]]
    header_lc = [h.lower() for h in header]

    name_idx = None
    for j, h in enumerate(header_lc):
        if _re.match(r"^invest(ment|ments)\b", h or ""):
            name_idx = j
            break
    if name_idx is None:
        return {
            "ok": False,
            "investments": 0,
            "values": 0,
            "sheet": resolved_sheet,
            "note": "No Investment/Investments column found in header.",
        }

    date_cols = _detect_date_columns(values, header_row_idx, preferred_year=preferred_year)
    if not date_cols:
        return {
            "ok": False,
            "error": "Could not locate monthly date columns for investments grid.",
            "sheet": resolved_sheet,
        }

    if preferred_year is not None:
        date_cols = {j: d for j, d in date_cols.items() if d.year == preferred_year}
        if not date_cols:
            return {
                "ok": False,
                "error": f"No monthly columns resolved for year {preferred_year}.",
                "sheet": resolved_sheet,
            }

    body = values[header_row_idx + 1 :]

    inserted_vals = 0
    ensured: Dict[str, int] = {}
    ensured_order: List[str] = []

    STOP_AT_NAME = "portfolio total"
    blank_streak = 0

    for r in body:
        name = str(r[name_idx] if name_idx < len(r) else "").strip()
        clean_name = _clean_txt(name)

        if clean_name == STOP_AT_NAME:
            break

        if not name:
            blank_streak += 1
            if blank_streak > 20:
                break
            continue
        blank_streak = 0
        if clean_name in {"total", "grand total"}:
            continue

        if name not in ensured:
            inv = Investment.query.filter_by(name=name).first()
            if not inv:
                inv = Investment(name=name, color_hex=_ensure_color(len(ensured_order)))
                db.session.add(inv)
                db.session.flush()
            ensured[name] = int(inv.id)
            ensured_order.append(name)
        inv_id = ensured[name]

        for j0b, mdt in sorted(date_cols.items(), key=lambda x: (x[1].year, x[1].month)):
            v = r[j0b] if j0b < len(r) else None
            f = _to_float_cell(v)
            if math.isnan(f):
                continue

            row = PortfolioInvestmentValue.query.filter_by(
                investment_id=inv_id, as_of_date=mdt
            ).first()
            if row is None:
                row = PortfolioInvestmentValue(investment_id=inv_id, as_of_date=mdt)
            row.value = float(f)
            row.source = "valuation_sheet"
            row.source_id = source_id
            db.session.add(row)
            inserted_vals += 1

    db.session.commit()
    return {
        "ok": True,
        "investments": len(ensured),
        "values": inserted_vals,
        "sheet": resolved_sheet,
        "year": preferred_year,
    }


# ---------------------- ROUTE: upload + ingest (ALL SHEETS) ----------------------
@excel_bp.post("/upload_and_ingest")
def upload_and_ingest():
    """
    Form-data:
      - file: Excel (.xlsx/.xlsm/.xls)
      - sheet: optional worksheet name (used only to process that sheet *first*)

    Behaviour now:
      - Opens the workbook
      - Iterates through **all sheets** in the file
      - For each sheet:
          * classify as investment / balance / mixed / unknown
          * if balance-style -> _ingest_new_balance_sheet(...)
          * if investment-style -> _ingest_investments_table(...)
      - Aggregates all admin periods & per-sheet results into the JSON response.
    """
    f = request.files.get("file")
    if not f:
        return jsonify(error="No file uploaded"), 400

    sheet_param = (request.form.get("sheet") or request.args.get("sheet") or "").strip()
    filename = secure_filename(f.filename or "")
    if not filename:
        return jsonify(error="Invalid filename"), 400

    # robust year parse (works with underscores/dashes/spaces) from file name
    m = re.search(r"(?<!\d)(20\d{2})(?!\d)", filename)
    preferred_year = int(m.group(1)) if m else None

    ext = os.path.splitext(filename)[1].lower()
    if ext not in {".xlsx", ".xlsm", ".xls"}:
        return jsonify(error=f"Unsupported type: {ext}"), 400

    path = os.path.join(UPLOAD_FOLDER, filename)
    f.save(path)

    # Optional lineage record (per-upload, not per-sheet)
    ds_id = None
    if DataSource is not None:
        try:
            added_by = None
            if current_user and getattr(current_user, "is_authenticated", False):
                added_by = getattr(current_user, "email", None) or getattr(
                    current_user, "username", None
                )
            ds = DataSource(
                kind="upload", file_name=filename, sheet_name=None, added_by=added_by
            )
            db.session.add(ds)
            db.session.flush()
            ds_id = ds.id
            db.session.commit()
        except Exception:
            db.session.rollback()
            ds_id = None  # non-fatal

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            sheet_names = list(wb.sheetnames or [])

            # Sheet processing order:
            # - if a specific sheet was requested, process that FIRST
            # - then process all remaining sheets
            ordered_sheets: List[str] = []
            if sheet_param:
                if sheet_param in sheet_names:
                    ordered_sheets.append(sheet_param)
                else:
                    # fuzzy/normalized matching
                    for sn in (_sheet_candidates(sheet_param) or [sheet_param]):
                        if sn in sheet_names:
                            ordered_sheets.append(sn)
                            break
                    if not ordered_sheets:
                        want = _normalize_sheet_name(sheet_param)
                        for n in sheet_names:
                            if _normalize_sheet_name(n) == want:
                                ordered_sheets.append(n)
                                break
                        if not ordered_sheets:
                            for n in sheet_names:
                                if want in _normalize_sheet_name(n):
                                    ordered_sheets.append(n)
                                    break
            for n in sheet_names:
                if n not in ordered_sheets:
                    ordered_sheets.append(n)

            # Aggregated results across all sheets
            admin_periods_all: set[str] = set()
            balance_results: List[Dict[str, Any]] = []
            investments_results: List[Dict[str, Any]] = []
            file_types_by_sheet: Dict[str, str] = {}
            first_processed_sheet: Optional[str] = None

            for resolved_sheet in ordered_sheets:
                ws = wb[resolved_sheet]
                values = [
                    [c.value for c in row]
                    for row in ws.iter_rows(
                        min_row=1, max_row=ws.max_row, max_col=ws.max_column
                    )
                ]

                # Skip completely empty sheets
                if not any(any(cell is not None for cell in row) for row in values):
                    continue

                normalized = _normalize_sheet_name(resolved_sheet)
                file_type = _classify_workbook(values or [])
                file_types_by_sheet[resolved_sheet] = file_type

                # NEW: treat 'master', 'investment', or 'investments' sheets as
                # investment tabs for the Investment database table.
                is_master = normalized == "master"
                is_invest_sheet = normalized in {"investment", "investments"}

                looks_balance = ("bcas" in normalized) or ("q4adj" in normalized) or (
                    file_type == "balance"
                )
                looks_invest = is_master or is_invest_sheet or (file_type == "investment")

                if file_type == "mixed":
                    # If a sheet is mixed, treat 'Master' or 'Investment' as investments
                    # and others as balance
                    looks_invest = is_master or is_invest_sheet
                    looks_balance = not looks_invest

                if first_processed_sheet is None:
                    first_processed_sheet = resolved_sheet

                # --- Balance-sheet ingest (Clarus Balance Sheet tabs) ---
                if looks_balance:
                    try:
                        bal_res = _ingest_new_balance_sheet(values or [], resolved_sheet)
                        balance_results.append(bal_res)
                        for p in bal_res.get("admin_periods_upserted", []) or []:
                            admin_periods_all.add(p)
                    except Exception as e:
                        traceback.print_exc()
                        balance_results.append(
                            {
                                "ok": False,
                                "sheet": resolved_sheet,
                                "error": f"Balance ingest failed: {e}",
                            }
                        )

                # --- Investment ingest (Master/Investment tab) ---
                if looks_invest:
                    try:
                        inv_res = _ingest_investments_table(
                            values or [], resolved_sheet, ds_id, preferred_year=preferred_year
                        )
                        investments_results.append(inv_res)
                    except Exception as e:
                        traceback.print_exc()
                        investments_results.append(
                            {
                                "ok": False,
                                "sheet": resolved_sheet,
                                "error": f"Investments ingest failed: {e}",
                            }
                        )

        finally:
            wb.close()

        # --- Record upload history & cleanup ---
        db.session.add(
            ExcelUploadHistory(filename=filename, uploaded_at=datetime.utcnow())
        )
        db.session.commit()
        try:
            os.remove(path)
        except Exception:
            pass

        # Decide top-level file_type value (for backwards compatibility)
        distinct_types = {t for t in file_types_by_sheet.values() if t}
        if len(distinct_types) == 1:
            top_file_type = next(iter(distinct_types))
        elif distinct_types:
            top_file_type = "multi"
        else:
            top_file_type = "unknown"

        return jsonify(
            {
                "ok": True,
                "sheet": first_processed_sheet,
                "sheets_processed": list(file_types_by_sheet.keys()),
                "file_type": top_file_type,
                "file_type_by_sheet": file_types_by_sheet,
                "data_source_id": int(ds_id) if ds_id is not None else None,
                "admin_periods_upserted": sorted(admin_periods_all),
                "investor_periods_result": balance_results,
                "investments_result": investments_results,
            }
        ), 200

    except (RuntimeError, SQLAlchemyError) as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify(error=f"Upload/ingest failed: {e}"), 500
    except Exception:
        db.session.rollback()
        traceback.print_exc()
        return jsonify(
            error="Upload/ingest failed. See server logs."
        ), 500
