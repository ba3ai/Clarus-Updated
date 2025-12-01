# routes/investor_sync_routes.py
from __future__ import annotations

import math, re, os
from datetime import datetime, date
from typing import Dict, Any, List, Tuple, Optional

import pandas as pd
from flask import Blueprint, jsonify, request
from flask_jwt_extended import jwt_required

from backend.extensions import db
from backend.models import SharePointConnection
# Prefer snapshot models; fall back to snapshot models file if needed
try:
    from backend.models import WorkbookSnapshot, InvestorBalance, InvestorPeriodBalance
except Exception:
    from backend.models_snapshot import WorkbookSnapshot, InvestorBalance, InvestorPeriodBalance  # type: ignore

# ─────────────────────────────────────────────────────────────────────────────
# Bearer helper (robust; same behavior as metrics sync)
# ─────────────────────────────────────────────────────────────────────────────
try:
    from backend.routes.sharepoint_excel_routes import _bearer_from_request
except Exception:
    _bearer_from_request = None

def _get_bearer(allow_session: bool = True) -> Optional[str]:
    from flask import has_request_context
    env_tok = (os.getenv("MS_GRAPH_BEARER") or "").strip()
    if env_tok:
        return env_tok
    if allow_session and has_request_context():
        auth = (request.headers.get("Authorization") or "").strip()
        if auth.lower().startswith("bearer "):
            return auth[7:].strip()
        if request.is_json:
            b = (request.json or {}).get("bearer")
            if b:
                return str(b).strip()
        if _bearer_from_request:
            try:
                b = _bearer_from_request()
                if b:
                    return b
            except Exception:
                pass
    return None

# ─────────────────────────────────────────────────────────────────────────────
# Parsing helpers
# ─────────────────────────────────────────────────────────────────────────────
def _to_num(x):
    if x is None:
        return math.nan
    if isinstance(x, (int, float)):
        try: return float(x)
        except Exception: return math.nan
    s = str(x).strip().replace(",", "").replace("$", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try: return float(s)
    except Exception: return math.nan

def _clean_txt(x: str) -> str:
    import regex as _re
    return _re.sub(r"\s+", " ", (x or "")).strip().lower()

def _good_display_name(s: str) -> bool:
    if not s: return False
    t = s.strip()
    if not t: return False
    # ignore bare ids
    if re.fullmatch(r"[A-Za-z0-9\-\_\/\.]+", t) and sum(ch.isalpha() for ch in t) < sum(ch.isdigit() for ch in t):
        return False
    return any(ch.isalpha() for ch in t) or (" " in t and len(t) >= 2)

def _find_header_row(values: List[List]) -> int:
    header_row_idx = 0
    for i in range(min(20, len(values))):
        row = [str(c).strip().lower() if c is not None else "" for c in values[i]]
        if any(("ending" in c and "balance" in c) for c in row):
            return i
    return header_row_idx

def _maybe_excel_serial(v):
    try:
        f = float(v)
        if 20000 < f < 90000:
            from datetime import datetime, timedelta
            return (datetime(1899, 12, 30) + timedelta(days=int(f))).date()
    except Exception:
        pass
    return None

def _parse_header_date(v) -> Optional[date]:
    from datetime import datetime
    if isinstance(v, datetime): return v.date()
    ser = _maybe_excel_serial(v)
    if ser: return ser
    try:
        d = pd.to_datetime(str(v).strip(), errors="coerce")
        if pd.notnull(d): return d.date()
    except Exception:
        pass
    return None

def _header_date_for_col(values: List[List], header_row_idx: int, col_idx: int) -> Optional[date]:
    r = header_row_idx - 1
    while r >= 0:
        row = values[r] if r < len(values) else []
        cell = row[col_idx] if col_idx < len(row) else None
        dt = _parse_header_date(cell)
        if dt: return dt
        r -= 1
    return None

def _parse_date_any(v) -> Optional[date]:
    from datetime import datetime
    if isinstance(v, datetime): return v.date()
    ser = _maybe_excel_serial(v)
    if ser: return ser
    try:
        d = pd.to_datetime(str(v).strip(), errors="coerce")
        if pd.notnull(d): return d.date()
    except Exception:
        pass
    return None

def _find_period_dates_row_map(values: List[List], max_scan_rows: int = 80) -> Dict[int, date]:
    """Return {col_1b: date} for the row labeled 'Ending Date', else best row of dates."""
    if not values: return {}
    rows = len(values)
    cols = max((len(r) for r in values), default=0)

    def _row_dates(r_1b: int):
        row = values[r_1b - 1] if r_1b - 1 < len(values) else []
        out = {}
        for c in range(1, cols + 1):
            v = row[c - 1] if c - 1 < len(row) else None
            d = _parse_date_any(v)
            if d: out[c] = d
        return out

    # explicit label
    for r in range(1, min(rows, max_scan_rows) + 1):
        row = values[r - 1]
        if any(_clean_txt(str(cell)) == "ending date" for cell in row):
            dates = _row_dates(r)
            if len(dates) >= 2:
                return dates

    # best date row
    best = {}
    for r in range(1, min(rows, max_scan_rows) + 1):
        dates = _row_dates(r)
        if len(dates) > len(best) and len(dates) >= 6:
            best = dates
    return best

# month END to match admin sync
def _month_end(d: date) -> date:
    from calendar import monthrange
    return date(d.year, d.month, monthrange(d.year, d.month)[1])

# ─────────────────────────────────────────────────────────────────────────────
# Column detection helpers
# ─────────────────────────────────────────────────────────────────────────────
def _idxs_from_headers(headers: List[str], pred_keywords: List[str]) -> List[int]:
    idxs = []
    for i, h in enumerate(headers):
        s = (h or "").lower().strip()
        if any(p in s for p in pred_keywords):
            idxs.append(i)
    return idxs

def _detect_metric_columns(values: List[List], header_row_idx: int) -> Dict[str, List[int]]:
    headers = [("" if h is None else str(h).strip()) for h in values[header_row_idx]]

    ending = _idxs_from_headers(headers, ["ending balance", "ending bal", "end balance", "current value"])
    unreal = _idxs_from_headers(headers, ["unrealized", "unrealised", "unrealized gain", "unrealized g/l", "unrealized pnl"])
    fees   = _idxs_from_headers(headers, ["management fee", "mgmt fee", "management fees"])

    # NEW: Operating expenses synonyms
    opex   = _idxs_from_headers(headers, [
        "operating expense", "operating expenses", "opex", "operating exp", "operating costs", "operational expenses"
    ])

    return {
        "ending": ending,
        "unrealized": unreal,
        "management_fees": fees,
        "operating_expenses": opex,   # NEW
    }


def _date_map_for_columns(values: List[List], header_row_idx: int, candidate_idxs: List[int]) -> Dict[int, date]:
    """
    Build a map {col_idx_0b -> period_month_end} for the provided columns.
    Strategy:
      1) If an explicit/“best” date row exists, map by nearest date column.
      2) Else, walk upward in the header for that column.
    """
    dates_1b: Dict[int, date] = _find_period_dates_row_map(values)
    date_all0b: Dict[int, date] = {c1b - 1: _month_end(d) for c1b, d in dates_1b.items()}
    out: Dict[int, date] = {}
    for j0b in candidate_idxs:
        d = date_all0b.get(j0b)
        if not d and date_all0b:
            nearest = min(date_all0b.keys(), key=lambda k: abs(k - j0b))
            if abs(nearest - j0b) <= 12:
                d = date_all0b[nearest]
        if not d:
            dt = _header_date_for_col(values, header_row_idx, j0b)
            if dt:
                d = _month_end(dt)
        if d:
            out[j0b] = d
    return out

# ─────────────────────────────────────────────────────────────────────────────
# Extract per-investor series for Ending, Unrealized, Mgmt Fees
# ─────────────────────────────────────────────────────────────────────────────
def _extract_investor_series(values: List[List], sheet: str) -> Tuple[
    Dict[str, List[Tuple[date, Optional[float]]]],       # ending_series
    Dict[str, List[Tuple[date, Optional[float]]]],       # unrealized_series
    Dict[str, List[Tuple[date, Optional[float]]]],       # mgmt_fees_series
    Dict[str, List[Tuple[date, Optional[float]]]],   # opex_series  <-- NEW
]:
    """
    Returns three dicts keyed by investor name with per-month series for:
    - Ending Balance
    - Unrealized Gain/Loss (if present)
    - Management Fees (if present)
    """
    if not values or len(values) < 2:
        return {}, {}, {}

    header_row_idx = _find_header_row(values)
    headers = [("" if h is None else str(h).strip()) for h in values[header_row_idx]]
    body = values[header_row_idx + 1:]
    if not body:
        return {}, {}, {}

    df = pd.DataFrame(body, columns=headers)

    # detect columns for each metric and map dates
    metric_cols = _detect_metric_columns(values, header_row_idx)
    ending_dates = _date_map_for_columns(values, header_row_idx, metric_cols["ending"])
    unreal_dates = _date_map_for_columns(values, header_row_idx, metric_cols["unrealized"])
    fees_dates   = _date_map_for_columns(values, header_row_idx, metric_cols["management_fees"])
    opex_dates   = _date_map_for_columns(values, header_row_idx, metric_cols["operating_expenses"])  # NEW

    # locate investor name column(s)
    def _idxs(pred_keywords):
        return _idxs_from_headers(headers, pred_keywords)
    name_kw = [
        "investor name", "investor", "limited partner", "lp name", "client name",
        "client", "partner", "account name", "holder", "entity name", "name",
    ]
    prefer_name_idxs = [i for i in _idxs(name_kw) if "id" not in (headers[i] or "").lower()]
    allow_id_name_idxs = [i for i in _idxs(name_kw) if "id" in (headers[i] or "").lower()]

    total_re = re.compile(r"\btotal\b", flags=re.I)
    label_width = min(25, df.shape[1])

    ending_series: Dict[str, List[Tuple[date, Optional[float]]]] = {}
    unreal_series: Dict[str, List[Tuple[date, Optional[float]]]] = {}
    fees_series:   Dict[str, List[Tuple[date, Optional[float]]]] = {}
    opex_series:   Dict[str, List[Tuple[date, Optional[float]]]] = {}   # NEW

    for _, row in df.iterrows():
        labels = [str(v or "").strip() for v in row.iloc[:label_width].tolist()]
        if not any(labels):
            continue
        if any(total_re.search(l) for l in labels):
            continue

        # Investor name resolution
        investor_name = None
        for i in prefer_name_idxs:
            val = str(row.iloc[i] or "").strip()
            if _good_display_name(val):
                investor_name = val; break
        if not investor_name:
            for i in allow_id_name_idxs:
                val = str(row.iloc[i] or "").strip()
                if _good_display_name(val):
                    investor_name = val; break
        if not investor_name:
            for val in labels:
                if val and not total_re.search(val) and _good_display_name(val):
                    investor_name = val; break
        if not investor_name:
            continue

        # Build the three series
        e_ser: List[Tuple[date, Optional[float]]] = []
        for j0b, mdt in sorted(ending_dates.items(), key=lambda x: (x[1].year, x[1].month)):
            v = row.iloc[j0b] if j0b < len(row) else None
            f = _to_num(v)
            e_ser.append((mdt, None if math.isnan(f) else float(f)))

        u_ser: List[Tuple[date, Optional[float]]] = []
        for j0b, mdt in sorted(unreal_dates.items(), key=lambda x: (x[1].year, x[1].month)):
            v = row.iloc[j0b] if j0b < len(row) else None
            f = _to_num(v)
            u_ser.append((mdt, None if math.isnan(f) else float(f)))

        f_ser: List[Tuple[date, Optional[float]]] = []
        for j0b, mdt in sorted(fees_dates.items(), key=lambda x: (x[1].year, x[1].month)):
            v = row.iloc[j0b] if j0b < len(row) else None
            f = _to_num(v)
            f_ser.append((mdt, None if math.isnan(f) else float(f)))

        # NEW: Operating expenses series
        x_ser = []
        for j0b, mdt in sorted(opex_dates.items(), key=lambda x: (x[1].year, x[1].month)):
            f = _to_num(row.iloc[j0b] if j0b < len(row) else None)
            x_ser.append((mdt, None if math.isnan(f) else float(f)))

        if not any(v is not None for _, v in e_ser):
            # we require at least ending balances to tie periods together
            continue

        if e_ser:
            ending_series[investor_name] = e_ser
        if u_ser:
            unreal_series[investor_name] = u_ser
        if f_ser:
            fees_series[investor_name] = f_ser
        if x_ser: opex_series[investor_name]   = x_ser   # NEW

    return ending_series, unreal_series, fees_series, opex_series

def _extract_rows_from_values(values: List[List], sheet: str) -> Tuple[List[dict], datetime | None, Optional[date], Optional[date]]:
    """Investor rollups (first/last EB), plus first/last dates for those columns (as month-end)."""
    if not values or len(values) < 2:
        return [], None, None, None

    header_row_idx = _find_header_row(values)
    headers = [("" if h is None else str(h).strip()) for h in values[header_row_idx]]
    body = values[header_row_idx + 1:]
    if not body:
        return [], None, None, None

    df = pd.DataFrame(body, columns=headers)

    def _idxs(pred_keywords):
        idxs = []
        for i, h in enumerate(headers):
            s = (h or "").lower().strip()
            if any(p in s for p in pred_keywords):
                idxs.append(i)
        return idxs

    eb_idxs = _idxs(["ending balance"]) or _idxs(["ending bal", "end balance"])
    if not eb_idxs:
        return [], None, None, None

    first_idx, last_idx = eb_idxs[0], eb_idxs[-1]
    first_col_date = _header_date_for_col(values, header_row_idx, first_idx)
    last_col_date  = _header_date_for_col(values, header_row_idx, last_idx)
    if first_col_date: first_col_date = _month_end(first_col_date)
    if last_col_date:  last_col_date  = _month_end(last_col_date)

    # 'as_of' aligns to last_col_date if no explicit date column
    as_of = None
    date_col_idxs = _idxs(["date", "month", "period"])
    if date_col_idxs:
        ser = pd.to_datetime(df.iloc[:, date_col_idxs[0]], errors="coerce")
        if ser.notna().any():
            as_of = pd.to_datetime(ser.max()).to_pydatetime()
    if not as_of and last_col_date:
        as_of = pd.to_datetime(last_col_date).to_pydatetime()

    name_kw = [
        "investor name", "investor", "limited partner", "lp name", "client name",
        "client", "partner", "account name", "holder", "entity name", "name",
    ]
    prefer_name_idxs = [i for i in _idxs(name_kw) if "id" not in (headers[i] or "").lower()]
    allow_id_name_idxs = [i for i in _idxs(name_kw) if "id" in (headers[i] or "").lower()]

    out = []
    total_re = re.compile(r"\btotal\b", flags=re.I)
    label_width = min(25, df.shape[1])

    for _, row in df.iterrows():
        labels = [str(v or "").strip() for v in row.iloc[:label_width].tolist()]
        if not any(labels):
            continue
        if any(total_re.search(l) for l in labels):
            continue

        first_v = _to_num(row.iloc[first_idx])
        last_v  = _to_num(row.iloc[last_idx])
        if math.isnan(first_v) and math.isnan(last_v):
            continue

        investor_name = None
        for i in prefer_name_idxs:
            val = str(row.iloc[i] or "").strip()
            if _good_display_name(val):
                investor_name = val; break
        if not investor_name:
            for i in allow_id_name_idxs:
                val = str(row.iloc[i] or "").strip()
                if _good_display_name(val):
                    investor_name = val; break
        if not investor_name:
            for val in labels:
                if val and not total_re.search(val) and _good_display_name(val):
                    investor_name = val; break
        if not investor_name:
            continue

        iv = 0.0 if math.isnan(first_v) else float(first_v)
        cv = 0.0 if math.isnan(last_v)  else float(last_v)
        moic = (cv / iv) if iv else None
        roi_pct = ((cv - iv) / iv * 100.0) if iv else None

        out.append({
            "investor": investor_name,
            "initial_value": iv,
            "current_value": cv,
            "initial_date": first_col_date,
            "current_date": last_col_date,
            "moic": moic,
            "roi_pct": roi_pct,
        })
    return out, as_of, first_col_date, last_col_date

# ─────────────────────────────────────────────────────────────────────────────
# Graph helpers
# ─────────────────────────────────────────────────────────────────────────────
from backend.graph_excel_live import create_session, close_session, used_range_values
from backend.graph_sharepoint import open_excel_by_share_url, list_worksheets

def _normalize_sheet_name(s: str) -> str:
    return re.sub(r"[\s\-\_\.\(\)\[\]\{\}\+]+", "", (s or "").strip()).lower()

def _sheet_candidates(name: str) -> list[str]:
    base = (name or "").strip()
    if not base: return []
    variants = {base, base.replace("(", " (").replace("  ", " ").strip(),
                re.sub(r"\s*\(", " (", base).strip(), re.sub(r"\s+", "", base), base.replace("+", " ")}
    return [v for v in variants if v]

# ─────────────────────────────────────────────────────────────────────────────
# Shared ingest helper (used by both endpoints)
# ─────────────────────────────────────────────────────────────────────────────
def _ingest_investor_values(values, sheet_input, drive_id, item_id, source):
    # now expects FOUR series (ending, unrealized, mgmt fees, operating expenses)
    ending_map, unreal_map, fees_map, opex_map = _extract_investor_series(values, sheet_input)

    rows, as_of, first_col_date, last_col_date = _extract_rows_from_values(values, sheet_input)
    if not rows:
        return jsonify(error="No investor rows parsed from worksheet"), 400

    # Clear prior snapshot for this (file,sheet)
    if WorkbookSnapshot:
        (InvestorBalance.query
            .filter(InvestorBalance.snapshot_id.in_(
                db.session.query(WorkbookSnapshot.id)
                .filter_by(sheet=sheet_input, drive_id=drive_id, item_id=item_id)
            ))).delete(synchronize_session=False)
        WorkbookSnapshot.query.filter_by(sheet=sheet_input, drive_id=drive_id, item_id=item_id) \
                              .delete(synchronize_session=False)
        db.session.flush()

    # Create new snapshot
    snap = WorkbookSnapshot(
        source=source,
        drive_id=drive_id,
        item_id=item_id,
        sheet=sheet_input,
        as_of=as_of or datetime.utcnow()
    )
    db.session.add(snap)
    db.session.flush()

    # Write balances + monthly period rows (month END aligned)
    for r in rows:
        inv_name = r["investor"]
        iv = float(r.get("initial_value") or 0.0)
        cv = float(r.get("current_value") or 0.0)

        irr_pct = None
        try:
            if iv > 0.0 and cv > 0.0 and r.get("initial_date") and r.get("current_date"):
                years = (pd.to_datetime(r["current_date"]) - pd.to_datetime(r["initial_date"])).days / 365.25
                if years > 0:
                    ratio = cv / iv
                    if ratio > 0:
                        irr_pct = float((ratio ** (1.0 / years) - 1.0) * 100.0)
        except Exception:
            irr_pct = None

        inv_row = InvestorBalance(
            snapshot_id=snap.id,
            investor=inv_name,
            initial_value=iv,
            current_value=cv,
            initial_date=r.get("initial_date"),
            current_date=r.get("current_date"),
            moic=(cv / iv) if iv else None,
            roi_pct=((cv - iv) / iv * 100.0) if iv else None,
            irr_pct=irr_pct,
        )
        db.session.add(inv_row)
        db.session.flush()

        # Per-month series for this investor
        series     = ending_map.get(inv_name, [])
        unreal_ser = {d: v for d, v in (unreal_map.get(inv_name, []) or [])}
        fees_ser   = {d: v for d, v in (fees_map.get(inv_name, [])   or [])}
        opex_ser   = {d: v for d, v in (opex_map.get(inv_name, [])   or [])}  # NEW

        prev_end = None
        for mdt, end_val in series:
            beginning = prev_end
            row = InvestorPeriodBalance.query.filter_by(investor=inv_name, period_date=mdt).first()
            if row is None:
                row = InvestorPeriodBalance(investor=inv_name, period_date=mdt)

            row.beginning_balance = beginning
            if end_val is not None:
                row.ending_balance = end_val

            # Optional per-month metrics when present in sheet
            if mdt in unreal_ser and unreal_ser[mdt] is not None:
                row.unrealized_gain_loss = unreal_ser[mdt]
            if mdt in fees_ser and fees_ser[mdt] is not None:
                row.management_fees = fees_ser[mdt]
            if mdt in opex_ser and opex_ser[mdt] is not None:                 # NEW
                row.operating_expenses = opex_ser[mdt]                         # NEW

            row.investor_balance_id = inv_row.id
            row.source = "sync"
            db.session.add(row)

            if end_val is not None:
                prev_end = end_val

    db.session.commit()
    return jsonify(
        ok=True,
        snapshot_id=snap.id,
        investors=len(rows),
        rows=len(rows),
        as_of=(as_of.isoformat() if as_of else None),
        first_date=(first_col_date.isoformat() if first_col_date else None),
        last_date=(last_col_date.isoformat() if last_col_date else None),
    ), 200

# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────
investor_sync_bp = Blueprint("investor_sync", __name__)

@investor_sync_bp.post("/api/investors/sync-workbook")
def sync_investors_from_workbook():
    """
    Body:
      { "url": "https://tenant.sharepoint.com/.../file.xlsx?...",
        "sheet": "bCAS (Q4 Adj)" }
      or
      { "upload_path": "/abs/path/file.xlsx", "sheet": "..." }
    """
    if not WorkbookSnapshot or not InvestorBalance:
        return jsonify(error="Snapshot models not available"), 500

    body = request.get_json(silent=True) or {}
    sheet_input = (body.get("sheet") or "bCAS (Q4 Adj)").strip()

    values = None
    drive_id = item_id = None
    source = None

    # A) SharePoint URL path
    url = (body.get("url") or "").strip()
    if url:
        bearer = _get_bearer()
        if not bearer:
            return jsonify(error="Missing Microsoft bearer (sign in or set MS_GRAPH_BEARER)."), 401
        drive_id, item_id = open_excel_by_share_url(url, bearer)
        sid = create_session(drive_id, item_id, bearer, persist=False)
        try:
            tried = []; last_err = None
            for sn in (_sheet_candidates(sheet_input) or [sheet_input]):
                tried.append(sn)
                try:
                    values = used_range_values(drive_id, item_id, sn, bearer, sid)
                    sheet_input = sn; break
                except Exception as e:
                    last_err = e
            if values is None:
                sheets = list_worksheets(drive_id, item_id, bearer, tenant_id=None) or []
                names = [s.get("name") if isinstance(s, dict) else str(s) for s in sheets]
                want = _normalize_sheet_name(sheet_input)
                resolved = None
                for n in names:
                    if _normalize_sheet_name(n) == want: resolved = n; break
                if not resolved:
                    for n in names:
                        if want in _normalize_sheet_name(n): resolved = n; break
                if resolved:
                    values = used_range_values(drive_id, item_id, resolved, bearer, sid)
                    sheet_input = resolved
                if values is None:
                    return jsonify(error=f"Worksheet not found. Tried {tried}. Last error: {last_err}"), 400
        finally:
            close_session(drive_id, item_id, bearer, sid)
        source = "sharepoint-live"

    # B) Local upload path
    upload_path = (body.get("upload_path") or "").strip()
    if (values is None) and upload_path:
        from openpyxl import load_workbook
        wb = load_workbook(upload_path, data_only=True, read_only=True)
        try:
            if sheet_input not in wb.sheetnames:
                cand = next((s for s in wb.sheetnames if s.strip().lower() == sheet_input.strip().lower()), None)
                if not cand:
                    cand = next((s for s in wb.sheetnames if sheet_input.strip().lower() in s.strip().lower()), None)
                if not cand:
                    return jsonify(error=f"Worksheet not found in upload: {sheet_input}"), 400
            ws = wb[sheet_input]
            values = [[c.value for c in row]
                      for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column)]
        finally:
            wb.close()
        source = "upload"

    if values is None:
        return jsonify(error="Provide either 'url' or 'upload_path'"), 400

    return _ingest_investor_values(values, sheet_input, drive_id, item_id, source or "upload")

@investor_sync_bp.post("/api/investors/sync")
@jwt_required(optional=True)
def sync_investors_from_connection():
    """Use the latest saved SharePoint connection; same ingest path as workbook."""
    body = request.get_json(silent=True) or {}
    sheet = (body.get("sheet") or "bCAS (Q4 Adj)").strip()

    conn = SharePointConnection.query.order_by(SharePointConnection.added_at.desc()).first()
    if not conn:
        return jsonify(error="No SharePoint connections saved."), 400

    bearer = _get_bearer()
    if not bearer:
        return jsonify(error="Missing Microsoft bearer (sign in or set MS_GRAPH_BEARER)."), 401

    sid = create_session(conn.drive_id, conn.item_id, bearer, persist=False)
    try:
        values = None; last_err = None; tried = []
        for sn in (_sheet_candidates(sheet) or [sheet]):
            tried.append(sn)
            try:
                values = used_range_values(conn.drive_id, conn.item_id, sn, bearer, sid)
                sheet = sn; break
            except Exception as e:
                last_err = e
        if values is None:
            sheets = list_worksheets(conn.drive_id, conn.item_id, bearer, tenant_id=None) or []
            names = [s.get("name") if isinstance(s, dict) else str(s) for s in sheets]
            want = _normalize_sheet_name(sheet)
            resolved = None
            for n in names:
                if _normalize_sheet_name(n) == want: resolved = n; break
            if not resolved:
                for n in names:
                    if want in _normalize_sheet_name(n): resolved = n; break
            if resolved:
                values = used_range_values(conn.drive_id, conn.item_id, resolved, bearer, sid)
                sheet = resolved
        if values is None:
            return jsonify(error=f"Worksheet not found. Tried {tried}. Last error: {last_err}"), 400
    finally:
        close_session(conn.drive_id, conn.item_id, bearer, sid)

    return _ingest_investor_values(values, sheet, conn.drive_id, conn.item_id, "sharepoint-live")
