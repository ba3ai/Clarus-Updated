# backend/services/excel_kpis.py
import io
from typing import Dict, Any, Optional
from openpyxl import load_workbook

def compute_metric(book_bytes: bytes, intent: Dict[str, Any], context: Dict[str, Any], meta: Dict[str, Any], user) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(book_bytes), data_only=True, read_only=True)

    action = intent.get("action")
    if action == "ending_balance_latest_total":
        sheet = context.get("sheet", "bCAS (Q4 Adj)")
        return _ending_balance_latest_total(wb, sheet, user)
    elif action == "ending_balance_total":
        sheet = context.get("sheet", "bCAS (Q4 Adj)")
        return _ending_balance_total(wb, sheet, user)
    elif action == "roi":
        return _roi_from_values(intent)
    elif action == "initial_value":
        sheet = context.get("sheet", "bCAS (Q4 Adj)")
        return _initial_value_guess(wb, sheet)
    elif action == "explain_formula":
        return {"type": "explanation", "answer": _explain(intent.get("topic"))}

    return {"type": "nlp", "answer": "I have your workbook. Ask for Current Value (latest Ending Balance), Ending Balance total, ROI, etc."}

def _col_index_by_header(ws, header_name: str) -> Optional[int]:
    for cell in ws[1]:
        val = (str(cell.value or "")).strip().lower()
        if val == header_name.strip().lower():
            return cell.col_idx
    return None

def _ending_balance_total(wb, sheet_name: str, user) -> Dict[str, Any]:
    ws = wb[sheet_name]
    col = _col_index_by_header(ws, "Ending Balance")
    if not col:
        for cell in ws[1]:
            text = str(cell.value or "").lower()
            if "ending" in text and "balance" in text:
                col = cell.col_idx
                break
    total = 0.0
    non_null = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        v = row[col-1].value if col else None
        if isinstance(v, (int, float)):
            total += float(v)
            non_null += 1
    return {"type": "metric", "metric": "ending_balance_total", "value": total, "count": non_null, "sheet": sheet_name}

def _ending_balance_latest_total(wb, sheet_name: str, user) -> Dict[str, Any]:
    ws = wb[sheet_name]
    candidates = ["date", "period", "month", "as of", "as-of", "period end"]
    date_col = None
    for cell in ws[1]:
        name = str(cell.value or "").strip().lower()
        if name in candidates:
            date_col = cell.col_idx
            break

    end_col = _col_index_by_header(ws, "Ending Balance")
    if not end_col:
        for cell in ws[1]:
            t = str(cell.value or "").lower()
            if "ending" in t and "balance" in t:
                end_col = cell.col_idx
                break

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        period = row[date_col-1].value if date_col else None
        val = row[end_col-1].value if end_col else None
        rows.append((period, val))

    last_period = None
    for period, _ in reversed(rows):
        if period not in (None, ""):
            last_period = period
            break

    if last_period is None:
        return _ending_balance_total(wb, sheet_name, user)

    total = 0.0
    count = 0
    for period, val in rows:
        if period == last_period and isinstance(val, (int, float)):
            total += float(val)
            count += 1

    return {
        "type": "metric",
        "metric": "ending_balance_latest_total",
        "period": str(last_period),
        "value": total,
        "count": count,
        "sheet": sheet_name,
    }

# Optional proxy for "initial value" if you don't store it separately:
def _initial_value_guess(wb, sheet_name: str) -> Dict[str, Any]:
    ws = wb[sheet_name]
    end_col = _col_index_by_header(ws, "Ending Balance")
    if not end_col:
        for cell in ws[1]:
            t = str(cell.value or "").lower()
            if "ending" in t and "balance" in t:
                end_col = cell.col_idx
                break
    first = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        v = row[end_col-1].value if end_col else None
        if isinstance(v, (int, float)):
            first = float(v)
            break
    return {"type": "metric", "metric": "initial_value", "value": first}

def _roi_from_values(intent: Dict[str, Any]) -> Dict[str, Any]:
    initial = intent.get("initial")
    current = intent.get("current")
    if initial in (None, 0):
        return {"type": "error", "message": "Initial value is required and must be non-zero."}
    roi = ((current - initial) / initial) * 100.0
    return {"type": "metric", "metric": "roi_pct", "value": roi}

def _explain(topic: Optional[str]) -> str:
    mapping = {
        "unrealized gain/loss": "Unrealized = Ending FV − (Beginning FV + Net Cash Flow). Sign indicates gain/loss without a sale.",
        "since inception": "Since inception aggregates from the fund start date up to the report as-of date (cumulative).",
    }
    key = (topic or "").lower()
    return mapping.get(key, "Ask about Ending Balance, Committed, Cash Flow, ROI, or provide a topic to explain.")
